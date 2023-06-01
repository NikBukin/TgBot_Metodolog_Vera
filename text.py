from openpyxl import load_workbook


## Токены стикеров в тг
sti_hi = 'CAACAgIAAxkBAAEHNCZjvlKt5wub29DEyV-uXNFCO7zgsQAC_B8AAsnOiEieiMmbEDXOyC0E'
sti_ok = 'CAACAgIAAxkBAAEHNCxjvlLZh3KlQ5FdzjM6EgT1OSWjigACpikAAmESgEiPzzOa3CbFyC0E'
sti_dream = 'CAACAgIAAxkBAAEHNCpjvlLMTvTjVXWCugxISJxMknOE-AACSB8AAkf0gEiW2c5LyC43cS0E'
sti_mail = 'CAACAgIAAxkBAAEHNCRjvlKcVWZFF0-w5Gwpjnmyq7KEUAACqx4AAhbvgEgyIYKQqXWvBC0E'
sti_done = 'CAACAgIAAxkBAAEHNC5jvlTelLhBDc5lUYy40NaqS8scqwACCyYAAqmIgEjdsXO0OoCZKy0E'
sti_love = 'CAACAgIAAxkBAAEHNChjvlK6QubUX8OHw8YGK1Y7e4VL0QACiyMAAvo2gUhCYGoFTq5QAy0E'


## Выгрузка текста из таблиц Excel
wb_Vera = load_workbook('alg_dir/Алгоритм Вера бот.xlsx')


sheet_full = wb_Vera['Общее']

start_1 = str(sheet_full['B2'].value)
start_2 = str(sheet_full['B4'].value)
start_3 = str(sheet_full['B6'].value)
finish_1 = str(sheet_full['B9'].value)
finish_sti = str(sheet_full['D9'].value)
finish_3 = str(sheet_full['F9'].value)
mark_grad = str(sheet_full['B11'].value)
grad_smile = str(sheet_full['B13'].value)
smile_list = [str(sheet_full['C16'].value), str(sheet_full['D16'].value), str(sheet_full['E16'].value)]
komment_1 = str(sheet_full['D19'].value)
komment_sti = str(sheet_full['G19'].value)
thank_for_grad = str(sheet_full['D21'].value)
thank_for_grad_sti = str(sheet_full['F21'].value)

## Проверка доступа
danger_1 = str(sheet_full['B24'].value)
danger_sti = str(sheet_full['D24'].value)
danger_2 = str(sheet_full['B26'].value)

danger_3 = str(sheet_full['B32'].value)
danger_4 = str(sheet_full['B34'].value)

done_check_mail = str(sheet_full['B38'].value)

sheet_DS = wb_Vera['ДС']


ds_start_text = str(sheet_DS['H6'].value)
ds_start_mark_neraz = str(sheet_DS['F9'].value)
ds_start_mark_priznak = str(sheet_DS['I9'].value)
ds_start_mark_uznat = str(sheet_DS['L9'].value)


DSone_def_neraz_text_1 = str(sheet_DS['F11'].value)
DSone_def_neraz_text_2 = str(sheet_DS['F13'].value)
DSone_def_neraz_mark_da = str(sheet_DS['F15'].value)
DSone_def_neraz_mark_net = str(sheet_DS['H15'].value)

DSone_def_priznak_text = str(sheet_DS['I11'].value)

DSone_def_uznat_text = str(sheet_DS['L11'].value)


DS_def_net_text = str(sheet_DS['H17'].value)
DS_def_net_mark_gpn = str(sheet_DS['H19'].value)
DS_def_net_mark_kc = str(sheet_DS['J19'].value)

DS_def_da_text = str(sheet_DS['F17'].value)


process_customer_step_def_text = str(sheet_DS['I21'].value)
process_customer_step_def_mark_rabota = str(sheet_DS['H23'].value)
process_customer_step_def_mark_mtr = str(sheet_DS['J23'].value)


process_subject_step_def_mtr_text = str(sheet_DS['L25'].value)
process_subject_step_def_mtr_mark_da = str(sheet_DS['K27'].value)
process_subject_step_def_mtr_mark_net = str(sheet_DS['P27'].value)

process_subject_step_def_rabota_text_1 = str(sheet_DS['F25'].value)
process_subject_step_def_rabota_text_2 = str(sheet_DS['C27'].value)
process_subject_step_def_rabota_mark_da = str(sheet_DS['B29'].value)
process_subject_step_def_rabota_mark_net = str(sheet_DS['D29'].value)


process_eol_step_def_text = process_subject_step_def_mtr_text
process_eol_step_def_mark_da = process_subject_step_def_mtr_mark_da
process_eol_step_def_mark_net = process_subject_step_def_mtr_mark_net


process_change_step_def_net_text = str(sheet_DS['R29'].value)
process_change_step_def_net_mark_1 = str(sheet_DS['Q31'].value)
process_change_step_def_net_mark_2 = str(sheet_DS['S31'].value)
process_change_step_def_net_mark_3 = str(sheet_DS['U31'].value)

process_change_step_def_da_text_1 = str(sheet_DS['J29'].value)
process_change_step_def_da_text_2 = str(sheet_DS['J31'].value)
process_change_step_def_da_mark_da = str(sheet_DS['G33'].value)
process_change_step_def_da_mark_net = str(sheet_DS['L33'].value)


process_changesize1_step_def_net_text = str(sheet_DS['L35'].value)
process_changesize1_step_def_net_mark_da = str(sheet_DS['K37'].value)
process_changesize1_step_def_net_mark_net = str(sheet_DS['M37'].value)


pers_1_1 = str(sheet_DS['Q33'].value)
pers_1_2 = str(sheet_DS['S33'].value)
pers_2 = str(sheet_DS['W37'].value)
pers_3 = str(sheet_DS['Z37'].value)
pers_4 = str(sheet_DS['AC37'].value)
pers_5 = str(sheet_DS['AF37'].value)
pers_6 = str(sheet_DS['AI37'].value)
pers_7 = str(sheet_DS['N46'].value)
pers_8 = str(sheet_DS['Q46'].value)
pers_9 = str(sheet_DS['T46'].value)
pers_10 = str(sheet_DS['W46'].value)
pers_11 = str(sheet_DS['D45'].value)
pers_12 = str(sheet_DS['G45'].value)
pers_13 = str(sheet_DS['J45'].value)
pers_14 = str(sheet_DS['B38'].value)
pers_15 = str(sheet_DS['E38'].value)
pers_16 = str(sheet_DS['H38'].value)
ds_temp_text = str(sheet_DS['H50'].value)



sheet_MSZ = wb_Vera['МСЗ']

msz_start = str(sheet_MSZ['J6'].value)
msz_start_mark1 = str(sheet_MSZ['I8'].value)
msz_start_mark2 = str(sheet_MSZ['K8'].value)

MSZ_def_prev = str(sheet_MSZ['L10'].value)
MSZ_def_ne_prev_1 = str(sheet_MSZ['I10'].value)
MSZ_def_ne_prev_2 = str(sheet_MSZ['I12'].value)
MSZ_def_ne_prev_mark1 = str(sheet_MSZ['H14'].value)
MSZ_def_ne_prev_mark2 = str(sheet_MSZ['J14'].value)


msz_step_one_def = str(sheet_MSZ['H16'].value)
msz_step_one_def_mark1 = str(sheet_MSZ['G18'].value)
msz_step_one_def_mark2 = str(sheet_MSZ['I18'].value)


msz_step_two_def_da = str(sheet_MSZ['E20'].value)
msz_step_two_def_da_mark1 = str(sheet_MSZ['D22'].value)
msz_step_two_def_da_mark2 = str(sheet_MSZ['F22'].value)

msz_step_two_def_net = str(sheet_MSZ['K20'].value)
msz_step_two_def_net_mark1 = str(sheet_MSZ['J22'].value)
msz_step_two_def_net_mark2 = str(sheet_MSZ['M22'].value)


msz_step_two_one_def_da = str(sheet_MSZ['C24'].value)
msz_step_two_one_def_net = str(sheet_MSZ['F24'].value)


msz_step_two_two_def_da = str(sheet_MSZ['I24'].value)
msz_step_two_two_def_da_mark1 = str(sheet_MSZ['I26'].value)
msz_step_two_two_def_da_mark2 = str(sheet_MSZ['L26'].value)

msz_step_two_two_def_net = str(sheet_MSZ['M24'].value)


msz_step_three_def_da = str(sheet_MSZ['I28'].value)
msz_step_three_def_net = str(sheet_MSZ['L28'].value)



sheet_EDP = wb_Vera['ЕдП']

EDP_start = str(sheet_EDP['L6'].value)
EDP_start_mark_1 = str(sheet_EDP['J8'].value)
EDP_start_mark_2 = str(sheet_EDP['N8'].value)


EDP_start_def_1 = str(sheet_EDP['J10'].value)
EDP_start_def_1_mark_da = str(sheet_EDP['I12'].value)
EDP_start_def_1_mark_net = str(sheet_EDP['K12'].value)

EDP_start_def_2 = str(sheet_EDP['H22'].value)
EDP_start_def_2_mark_da = str(sheet_EDP['G24'].value)
EDP_start_def_2_mark_net = str(sheet_EDP['I24'].value)
####

market_def_da = str(sheet_EDP['F14'].value)
market_def_da_mark_da = str(sheet_EDP['E16'].value)
market_def_da_mark_net = str(sheet_EDP['G16'].value)

market_def_net = str(sheet_EDP['J14'].value)
market_def_net_mark_da = str(sheet_EDP['I16'].value)
market_def_net_mark_net = str(sheet_EDP['K16'].value)


restr_def_net = str(sheet_EDP['K18'].value)

restr_def_da = market_def_net
restr_def_da_mark_da = market_def_net_mark_da
restr_def_da_mark_net = market_def_net_mark_net


biglist_def_net = str(sheet_EDP['K18'].value)

biglist_def_da = str(sheet_EDP['I18'].value)
biglist_def_da_mark_da = str(sheet_EDP['H20'].value)
biglist_def_da_mark_net = str(sheet_EDP['J20'].value)


hard_quest_def_da = EDP_start_def_2

hard_quest_def_da_mark_da = EDP_start_def_2_mark_da
hard_quest_def_da_mark_net = EDP_start_def_2_mark_net


summ_def = str(sheet_EDP['K26'].value)
summ_def_mark_kc = str(sheet_EDP['J28'].value)
summ_def_mark_gpn = str(sheet_EDP['L28'].value)


EDP_subject_def_gpn_net = str(sheet_EDP['H34'].value)
EDP_subject_def_gpn_net_mark_da = str(sheet_EDP['G36'].value)
EDP_subject_def_gpn_net_mark_net = str(sheet_EDP['I36'].value)

EDP_subject_def = str(sheet_EDP['H30'].value)
EDP_subject_def_mark_da = str(sheet_EDP['G32'].value)
EDP_subject_def_mark_net = str(sheet_EDP['I32'].value)


EDP_eol_def = EDP_subject_def_gpn_net
EDP_eol_def_mark_da = EDP_subject_def_gpn_net_mark_da
EDP_eol_def_mark_net = EDP_subject_def_gpn_net_mark_net

EDP_vers_0 = str(sheet_EDP['A38'].value)
EDP_vers_1 = str(sheet_EDP['D43'].value)
EDP_vers_2 = str(sheet_EDP['G43'].value)
EDP_vers_3 = str(sheet_EDP['J43'].value)
EDP_vers_4 = str(sheet_EDP['M43'].value)
EDP_vers_5 = str(sheet_EDP['P43'].value)
EDP_vers_6 = str(sheet_EDP['S43'].value)
EDP_vers_7 = str(sheet_EDP['V43'].value)
EDP_documents = str(sheet_EDP['I47'].value)
EDP_templates = str(sheet_EDP['I49'].value)








wb_quest = load_workbook('quest_dir/Алгоритм Вопрос-Ответ.xlsx')

sheet_quest = wb_quest['вопрос-ответ']

hi = str(sheet_quest['I2'].value)

mark_quest = str(sheet_quest['G5'].value)
quest = str(sheet_quest['G7'].value)
quest_Vera = str(sheet_quest['G9'].value)
quest_zero = str(sheet_quest['G13'].value)
quest_fin = str(sheet_quest['G15'].value)

mark_gloss = str(sheet_quest['M5'].value)
gloss = str(sheet_quest['M7'].value)
gloss_Vera = quest_Vera
gloss_zero = quest_zero
gloss_fin = str(sheet_quest['M15'].value)


mark_send = str(sheet_quest['E18'].value)
send_quest = str(sheet_quest['E20'].value)
